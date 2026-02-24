"""
================================================================================
HU Analyzer — Actinver | Área de Desarrollo de Productos Digitales
================================================================================
Analiza DEFINICIÓN FUNCIONAL y CAPAS TECNOLÓGICAS INVOLUCRADAS de HUs con IA.
La información técnica detallada se revisa en prerefinamiento y refinamiento.
Agrega columnas de score + brechas al Excel + hoja de Síntesis ejecutiva.

INSTALACIÓN:
    pip install anthropic openpyxl

CONFIGURACIÓN:
    Windows:   set ANTHROPIC_API_KEY=sk-ant-...
    Mac/Linux: export ANTHROPIC_API_KEY=sk-ant-...

USO:
    python hu_analyzer.py --input HUs_Compilado.xlsx
    python hu_analyzer.py --input HUs_Compilado.xlsx --output resultado.xlsx
    python hu_analyzer.py --input HUs_Compilado.xlsx --limit 5   # prueba rápida
    python hu_analyzer.py --input HUs_Compilado.xlsx --sheet "Onboarding"

SALIDA:
    - Por defecto: carpeta Output/ con archivos HUs_Compilado_analizado_v1.0.xlsx,
      v2.0.xlsx, v3.0.xlsx... (numeración consecutiva en cada ejecución)
    - Columnas de análisis + hoja "Síntesis Ejecutiva"
================================================================================
"""

import anthropic
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import json
import argparse
import sys
import os
import re
import time
import difflib
from concurrent.futures import ThreadPoolExecutor, as_completed
from datetime import datetime

class AnthropicGameOverError(Exception):
    """Se acabaron tokens/créditos en la cuenta Anthropic."""


# Carpeta donde se guardan los archivos de análisis (con numeración v1.0, v2.0...)
OUTPUT_DIR = "Output"

# Archivo para persistir HU Speed (promedio de segundos por HU)
_HU_SPEED_FILE = os.path.join(os.path.dirname(os.path.abspath(__file__)), ".hu_analyzer_speed.json")


def get_hu_speed() -> float | None:
    """Retorna el promedio de segundos por HU (HU Speed Analysis) o None si no hay datos."""
    try:
        if os.path.exists(_HU_SPEED_FILE):
            with open(_HU_SPEED_FILE) as f:
                data = json.load(f)
            return float(data.get("avg_seconds", 0)) or None
    except Exception:
        pass
    return None


def update_hu_speed(elapsed_sec: float) -> float:
    """
    Actualiza HU Speed con el tiempo de una HU analizada.
    Usa promedio móvil: nuevo_avg = (avg_anterior * n + elapsed) / (n + 1)
    Retorna el nuevo promedio.
    """
    try:
        data = {"avg_seconds": 0.0, "count": 0}
        if os.path.exists(_HU_SPEED_FILE):
            with open(_HU_SPEED_FILE) as f:
                data = json.load(f)
        avg = float(data.get("avg_seconds", 0))
        n = int(data.get("count", 0))
        new_avg = (avg * n + elapsed_sec) / (n + 1)
        with open(_HU_SPEED_FILE, "w") as f:
            json.dump({"avg_seconds": round(new_avg, 2), "count": n + 1}, f)
        return new_avg
    except Exception:
        return elapsed_sec


def count_hus_to_analyze(input_path: str, target_sheet: str = None, limit: int = None) -> int:
    """Cuenta cuántas HUs se analizarán (análisis de alto nivel, rápido)."""
    _, all_hus = load_all_hus(input_path, target_sheet, quiet=True)
    total = len(all_hus)
    return min(total, limit) if limit else total


try:
    from config import MAX_CONCURRENT_ANALYSIS
except ImportError:
    MAX_CONCURRENT_ANALYSIS = 5


# ══════════════════════════════════════════════════════════════════════════════
# 1. CONSTANTES DE ESTRUCTURA DEL EXCEL
#    Ajusta estos valores si tu Excel tiene diferente layout
# ══════════════════════════════════════════════════════════════════════════════

HEADER_ROW    = 8    # Fila donde están los encabezados de columnas
DATA_START_ROW = 9   # Primera fila de datos reales
SKIP_VALUES = {"Ejemplo", "EJEMPLO", "ejemplo", ""}  # Filas a ignorar por ID


# ══════════════════════════════════════════════════════════════════════════════
# 2. DIMENSIONES DE EVALUACIÓN Y PESOS
# ══════════════════════════════════════════════════════════════════════════════

DIMENSIONS = {
    "funcional": "Definición Funcional",
    "capas_tec": "Capas Tecnológicas Involucradas",
    "ux_ui":     "UX / UI (funcional)",
    "integraciones": "Integraciones/Sistemas",
    "regulatorio": "Regulatorio & Seguridad",
    "criterios":   "Criterios de Aceptación",
}

# Pesos: enfoque en definición funcional. Lo técnico se revisa en prerefinamiento/refinamiento.
DIMENSION_WEIGHTS = {
    "funcional":      0.35,
    "capas_tec":      0.25,
    "ux_ui":          0.15,
    "integraciones":  0.10,
    "regulatorio":    0.08,
    "criterios":      0.07,
}

# Columnas de análisis: (texto header, color fondo hex, ancho columna)
ANALYSIS_HEADERS = [
    ("SCORE\nTOTAL\n(0-100)",                   "1F3864", 9),
    ("NIVEL DE\nCOMPLETITUD",                    "1F3864", 14),
    ("SCORE\nFuncional\n(0-10)",                 "2E75B6", 9),
    ("SCORE\nCapas Tec.\n(0-10)",                "2E75B6", 9),
    ("SCORE\nUX/UI\n(0-10)",                     "2E75B6", 9),
    ("SCORE\nIntegr.\n(0-10)",                   "2E75B6", 9),
    ("SCORE\nRegulat.\n(0-10)",                  "2E75B6", 9),
    ("SCORE\nCriterios\n(0-10)",                 "2E75B6", 9),
    ("CAPAS TECNOLÓGICAS\nINVOLUCRADAS",        "44546A", 35),
    ("RESUMEN\nEJECUTIVO",                       "375623", 45),
    ("POR DEFINIR\nFuncional",                   "843C0C", 38),
    ("POR DEFINIR\nCapas Tec.",                  "843C0C", 38),
    ("POR DEFINIR\nUX / UI",                     "843C0C", 38),
    ("POR DEFINIR\nIntegraciones",               "843C0C", 38),
    ("POR DEFINIR\nRegulatorio",                  "843C0C", 38),
    ("POR DEFINIR\nCriterios",                   "843C0C", 38),
    ("PREGUNTAS PARA\nCLARIFICAR",               "7B2C2C", 45),
    ("MEJORAS\nIDENTIFICADAS",                   "375623", 40),
    ("COMPARACIÓN\nvs ANTERIOR",                 "44546A", 35),
]


# ══════════════════════════════════════════════════════════════════════════════
# 3. PROMPTS PARA CLAUDE
# ══════════════════════════════════════════════════════════════════════════════

SYSTEM_PROMPT = """Eres un Product Manager Senior con 15 años en banca digital mexicana.
Evalúas Historias de Usuario del área de Productos Digitales de Actinver.

CONTEXTO CLAVE: Las HUs son definidas por personas de NEGOCIO (POs, Product Owners).
La parte técnica NO está involucrada aún — eso ocurre en prerefinamiento y refinamiento.
Tu rol es apoyar al PO a definir funcionalmente la iniciativa antes de que llegue a equipos técnicos.

ESTRUCTURA DE LAS HUs: Cada fila es una ETAPA del proceso de la iniciativa. Una sola HU
cubre solo una parte del flujo completo; el flujo completo se arma con TODAS las HUs.
Por tanto: NO esperes toda la información en cada HU. Una HU inicial puede no mencionar
regulatorio porque eso se define en una etapa posterior; una HU de validación puede no
tener criterios de aceptación detallados si eso está en otra etapa. Evalúa qué es
razonable definir EN ESA ETAPA del proceso, no qué debería tener el flujo completo.

REGULATORIO: Puede aparecer de dos formas: (a) en una etapa específica donde aplica, o
(b) en una HU dedicada que cubre TODA la parte regulatoria de la iniciativa. Si una HU
hace referencia a toda la parte regulatoria (CUB, PLD/AML, LFPDPPP, etc.), valórala
positivamente: es una HU que concentra el alcance regulatorio de la iniciativa.

TONO: Sé constructivo y orientado a acción. En lugar de señalar "qué falta", indica
"qué conviene definir" para que la HU esté lista para prerefinamiento. Usa lenguaje
accesible para personas de negocio (evita jerga técnica).

IMPORTANTE: La DEFINICIÓN FUNCIONAL debe incluir explícitamente: (1) Mensajes de error,
(2) Flujos alternos, (3) Mecanismos de medición y monitoreo. Evalúa si el PO los definió.
También evalúa CAPAS TECNOLÓGICAS INVOLUCRADAS. No pidas specs técnicas — eso se define después.

Contexto Actinver:
- Core: Core Bancario | Integraciones: RENAPO, INE, SAT, Buró, SPEI, biométricos
- Notificaciones: push, SMS, email | Regulatorio: CUB, PUI, LFPDPPP, PLD/AML
- Productos: Onboarding N4, Cuenta Remunerada, Crédito Simple, Tarjeta Débito

Responde ÚNICAMENTE con JSON válido. Sin texto antes ni después del JSON."""


def build_analysis_prompt(hu: dict, prev_data: dict = None) -> str:
    hu_text = "\n".join(
        f"  {k}: {v}"
        for k, v in hu.items()
        if v and str(v).strip() not in ("nan", "None", "")
        and not k.startswith("_")
    )

    prev_block = ""
    if prev_data:
        brechas_prev = prev_data.get("brechas") or {}
        brechas_txt = " | ".join(f"{k}: {(str(v)[:60]+'...' if len(str(v))>60 else v)}" for k, v in brechas_prev.items() if v)
        prev_block = f"""
═══ ANÁLISIS ANTERIOR (referencia) ═══
Score total previo: {prev_data.get('score_total', 0):.0f}/100
Nivel previo: {prev_data.get('nivel', '')}
Resumen previo: {str(prev_data.get('resumen', ''))[:300]}
Brechas previas: {brechas_txt[:400]}
══════════════════════════════════════
COMPARA la HU actual con el análisis anterior. Ubica la HU (por ID) e identifica las MEJORAS que el PO hizo.
Normalmente el score debería subir. Si no sube o baja, indica en comparacion_anterior qué antes estaba mejor definido.
"""

    return f"""Analiza la siguiente Historia de Usuario de Actinver. Evalúa su
DEFINICIÓN FUNCIONAL desde la perspectiva del PO (persona de negocio).
La parte técnica se abordará en prerefinamiento — aquí solo lo funcional.

RECUERDA: Cada HU es una ETAPA del proceso. No esperes toda la información en cada una.
El flujo completo se arma con todas las HUs. Evalúa qué es razonable para ESTA etapa
(no penalices por info que puede estar en otra etapa posterior o anterior).

═══ HISTORIA DE USUARIO ═══
{hu_text}
═══════════════════════════
{prev_block}
Evalúa las 6 dimensiones (score 0-10) desde lo que el PO debe definir:

1. DEFINICIÓN FUNCIONAL (35%): Debe incluir explícitamente: (a) Mensajes de error,
   (b) Flujos alternos, (c) Mecanismos de medición y monitoreo. Además: flujo principal,
   reglas de negocio, casos límite. ¿Está claro qué hace el sistema?

2. CAPAS TECNOLÓGICAS INVOLUCRADAS (25%): ¿El PO identificó qué capas toca?
   (UI, Backend, Integraciones como RENAPO/SAT/Core Bancario, Seguridad, Notificaciones)
   Solo identificación — no specs técnicas.

3. UX/UI FUNCIONAL (15%): ¿Qué debe definir el PO? Estados de pantalla (cargando,
   vacío, error, éxito), validaciones, flujo de navegación, feedback al usuario.

4. INTEGRACIONES/SISTEMAS (10%): ¿El PO identificó qué sistemas intervienen?
   (RENAPO, SAT, Buró, SPEI, Core Bancario, etc.) Solo identificación funcional.

5. REGULATORIO & SEGURIDAD (8%): ¿El PO identificó aspectos regulatorios?
   (CUB, PLD/AML, LFPDPPP, datos sensibles) Puede ser: (a) en esta etapa si aplica,
   (b) una HU dedicada que cubre TODA la parte regulatoria de la iniciativa. Si la HU
   hace referencia a toda la parte regulatoria, valórala positivamente.

6. CRITERIOS DE ACEPTACIÓN (7%): ¿El PO definió criterios testeables y medibles?
   Solo si aplica a esta etapa. No exigir en cada HU si el flujo completo los tiene.

Responde SOLO con este JSON exacto (sin markdown, sin bloques ```, sin texto extra):
{{
  "scores": {{
    "funcional": <0-10>,
    "capas_tec": <0-10>,
    "ux_ui": <0-10>,
    "integraciones": <0-10>,
    "regulatorio": <0-10>,
    "criterios": <0-10>
  }},
  "capas_tecnologicas": "<lista de capas involucradas separadas por | ej: UI | Backend | RENAPO | Notificaciones>",
  "resumen": "<2 oraciones constructivas: nivel de definición funcional y qué falta definir para prerefinamiento>",
  "brechas": {{
    "funcional":  "<qué conviene que el PO defina en esta dimensión | o 'Completo' si está listo>",
    "capas_tec":  "<qué capas conviene que el PO identifique | o 'Completo'>",
    "ux_ui":      "<qué conviene definir en UX/UI | o 'Completo'>",
    "integraciones": "<qué sistemas conviene identificar | o 'Completo'>",
    "regulatorio": "<qué aspectos regulatorios conviene identificar | o 'Completo'>",
    "criterios":   "<qué criterios conviene definir | o 'Completo'>"
  }},
  "preguntas_criticas": "<3-5 preguntas amigables para que el PO clarifique antes del prerefinamiento, separadas por |>",
  "mejoras_identificadas": "<si hay análisis anterior: mejoras que el PO hizo a la HU, separadas por |. Si no hay anterior: 'N/A'>",
  "comparacion_anterior": "<si hay análisis anterior y el score bajó: 'Anteriormente estaba mejor definido en: [aspectos]'. Si subió o no hay anterior: 'N/A'>"
}}

En "brechas": escribe QUÉ DEBE DEFINIR el PO para esta etapa, en tono constructivo.
No pidas info que corresponda a otras etapas del flujo. Si para esta etapa está completo: "Completo".
Si hay análisis anterior: en "mejoras_identificadas" lista las mejoras; en "comparacion_anterior" solo escribe algo si el score BAJÓ respecto al anterior."""


EXECUTIVE_ANALYSIS_PROMPT = """Eres el analista de HUs de Productos Digitales Actinver. Te encuentras en el elevador con un líder que te pregunta: "¿Cómo van las iniciativas de cada PO?"

Tienes el resumen del análisis de HUs por iniciativa. Para CADA iniciativa debes escribir UN párrafo corto (3-5 oraciones) que:
1. Resuma cómo va la iniciativa en general (nivel de definición, fortalezas).
2. Indique qué se mejoró en las HUs (si hay mejoras identificadas o comparación con análisis anterior).
3. Señale errores, inconsistencias o problemas en la información/data si los detectas (ej: IDs duplicados, datos faltantes, contradicciones, HUs mal estructuradas).
4. Sea directo y ejecutivo — como si respondieras en 30 segundos en el elevador.

Tono: profesional, constructivo, orientado a acción. Si todo va bien, dilo. Si hay riesgos o gaps, dilo sin dramatizar.
Responde ÚNICAMENTE con JSON válido. Sin texto antes ni después.

Formato de respuesta:
{
  "iniciativas": [
    {
      "nombre": "<nombre exacto de la iniciativa>",
      "analisis_ejecutivo": "<párrafo completo para esta iniciativa>"
    }
  ]
}"""


# ══════════════════════════════════════════════════════════════════════════════
# 4. OUTPUT CON VERSIONADO
# ══════════════════════════════════════════════════════════════════════════════

def get_next_output_path(input_path: str, original_filename: str = None) -> str:
    """
    Genera ruta en Output/ con numeración consecutiva v1.0, v2.0, v3.0...
    original_filename: nombre original para upload (ej: HUs_Compilado.xlsx)
    """
    script_dir = os.path.dirname(os.path.abspath(__file__))
    output_dir = os.path.join(script_dir, OUTPUT_DIR)
    os.makedirs(output_dir, exist_ok=True)

    if original_filename:
        base = os.path.splitext(os.path.basename(original_filename))[0]
        ext = os.path.splitext(original_filename)[1] or ".xlsx"
    else:
        base = os.path.splitext(os.path.basename(input_path))[0]
        ext = os.path.splitext(input_path)[1] or ".xlsx"
    pattern = re.compile(rf"^{re.escape(base)}_analizado_v(\d+)\.0{re.escape(ext)}$")

    max_v = 0
    if os.path.isdir(output_dir):
        for f in os.listdir(output_dir):
            m = pattern.match(f)
            if m:
                max_v = max(max_v, int(m.group(1)))

    next_v = max_v + 1
    filename = f"{base}_analizado_v{next_v}.0{ext}"
    return os.path.join(output_dir, filename)


# ══════════════════════════════════════════════════════════════════════════════
# 5. LECTURA DEL EXCEL
# ══════════════════════════════════════════════════════════════════════════════

# Pistas para detectar fila de encabezados (hojas con estructura distinta)
HEADER_HINTS = ("id", "no. hu", "no hu", "titulo", "título", "descripción", "descripcion", "historia")

def _detect_header_row(ws) -> int:
    """Detecta la fila de encabezados escaneando las primeras 15 filas."""
    max_scan = min(16, getattr(ws, "max_row", 100) + 1)
    for row_idx in range(1, max_scan):
        try:
            row = ws[row_idx]
            cells = [str(c.value or "").strip().lower() for c in list(row)[:10]]
        except (TypeError, AttributeError):
            # read_only: usar iter_rows
            rows = list(ws.iter_rows(min_row=row_idx, max_row=row_idx, max_col=10))
            cells = [str(c.value or "").strip().lower() for c in rows[0]] if rows else []
        if any(h in " ".join(cells) for h in HEADER_HINTS):
            return row_idx
    return HEADER_ROW

def _find_id_column(headers: list[str]) -> int:
    """Encuentra el índice de la columna ID."""
    for i, h in enumerate(headers):
        h_lower = (h or "").strip().lower()
        if h_lower in ("id", "no. hu", "no hu", "hu-id", "hu id", "código", "codigo"):
            return i
    return 0


def _normalize_hu_id(hu_id: str) -> str:
    """
    Normaliza ID para matching flexible: HU-001, HU-1, 001, 1, 1.0 -> hu_1.
    Evita que diferencias de formato (Excel convierte números) impidan el match.
    """
    s = str(hu_id or "").strip()
    if not s:
        return ""
    # Excel puede devolver 1.0 en vez de "HU-001"
    if s.replace(".", "").isdigit():
        return f"hu_{int(float(s))}"
    m = re.search(r"HU[-_]?\s*(\d+)", s, re.I)
    if m:
        return f"hu_{int(m.group(1))}"
    m = re.search(r"^(\d+)", s)
    if m:
        return f"hu_{int(m.group(1))}"
    return s.lower()


def _normalize_title(title: str) -> str:
    """Normaliza título para matching: minúsculas, sin espacios extra."""
    s = str(title or "").strip()
    return " ".join(s.lower().split()) if s else ""


def _normalize_content(text: str, max_len: int = 300) -> str:
    """Normaliza descripción/contenido para comparación semántica."""
    s = str(text or "").strip()
    s = " ".join(s.lower().split())
    return s[:max_len] if s else ""


def _find_title_column(headers: list[str]) -> int:
    """Encuentra el índice de la columna Título."""
    for i, h in enumerate(headers):
        h_lower = (h or "").strip().lower()
        if h_lower in ("titulo", "título", "title"):
            return i
    return 1  # fallback: columna B típica


def _find_desc_column(headers: list[str]) -> int:
    """Encuentra el índice de la columna Descripción o Definición."""
    for i, h in enumerate(headers):
        h_lower = (h or "").strip().lower()
        if h_lower in ("descripcion", "descripción", "description", "definicion", "definición", "definición funcional"):
            return i
    return 2  # fallback: columna C típica


def _similarity(a: str, b: str) -> float:
    """Ratio de similitud 0-1 entre dos strings."""
    if not a or not b:
        return 0.0
    return difflib.SequenceMatcher(None, a, b).ratio()

def get_sheet_headers(ws) -> list[str]:
    """Obtiene la lista de encabezados de una hoja (para usar como formato común)."""
    header_row = _detect_header_row(ws)
    return [str(cell.value or "").strip() for cell in ws[header_row]]


def get_common_headers_from_excel(excel_path: str) -> list[str]:
    """
    Obtiene los encabezados comunes del Excel base (primera hoja con datos).
    Usado para que las hojas de Word coincidan con el formato de las iniciativas existentes.
    """
    wb = openpyxl.load_workbook(excel_path)
    for sheet_name in wb.sheetnames:
        if sheet_name == "📊 Síntesis Ejecutiva":
            continue
        ws = wb[sheet_name]
        headers = get_sheet_headers(ws)
        if headers and any(h for h in headers):
            return [h for h in headers if h]
    return []


def read_sheet_hus(ws, sheet_name: str) -> list[dict]:
    """Lee todas las HUs válidas de una hoja. Detecta header row si la hoja tiene estructura distinta."""
    header_row = _detect_header_row(ws)
    headers = [
        str(cell.value).strip() if cell.value else ""
        for cell in ws[header_row]
    ]
    id_col = _find_id_column(headers)
    data_start = header_row + 1

    hus = []
    for row_idx in range(data_start, ws.max_row + 1):
        row = ws[row_idx]
        hu_id = str(row[id_col].value).strip() if row[id_col].value else ""
        if hu_id in SKIP_VALUES or not hu_id:
            continue

        hu = {"_sheet": sheet_name, "_row": row_idx, "_hu_id": hu_id}
        for col_idx, header in enumerate(headers):
            if header:
                val = row[col_idx].value if col_idx < len(row) else None
                hu[header] = str(val).strip() if val else ""
        hus.append(hu)

    return hus


def load_all_hus(filepath: str, target_sheet: str = None, quiet: bool = False) -> tuple[dict, list]:
    """Carga HUs de todas las hojas (o solo la especificada)."""
    def _log(msg):
        if not quiet:
            print(msg)

    wb = openpyxl.load_workbook(filepath)
    sheets_data = {}
    all_hus = []

    sheets = [target_sheet] if target_sheet else wb.sheetnames
    for sheet_name in sheets:
        if sheet_name not in wb.sheetnames:
            _log(f"  ⚠  Hoja '{sheet_name}' no encontrada, se omite.")
            continue
        ws = wb[sheet_name]
        hus = read_sheet_hus(ws, sheet_name)
        sheets_data[sheet_name] = hus
        all_hus.extend(hus)
        hdr = _detect_header_row(ws)
        extra = f" (headers fila {hdr})" if hdr != HEADER_ROW else ""
        _log(f"  📋  {sheet_name}: {len(hus)} HUs cargadas{extra}")

    return sheets_data, all_hus


def load_previous_analysis(filepath: str, log_fn=None) -> dict:
    """
    Carga el análisis previo desde un Excel ya analizado.
    Retorna dict con:
      - "by_id": (sheet, norm_id) -> prev_data
      - "by_sheet": sheet -> list of {prev, norm_title, norm_desc}

    Matching: 1) ID, 2) título exacto, 3) título + contenido similar.
    """
    def _log(msg):
        if log_fn:
            log_fn(msg)

    prev_by_id: dict[tuple[str, str], dict] = {}
    prev_by_sheet: dict[str, list] = {}
    wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)

    for sheet_name in wb.sheetnames:
        if sheet_name == "📊 Síntesis Ejecutiva":
            continue
        ws = wb[sheet_name]
        header_row = _detect_header_row(ws)
        headers = [str(c.value or "").strip() for c in ws[header_row]]
        id_col = _find_id_column(headers)
        title_col = _find_title_column(headers)
        desc_col = _find_desc_column(headers)
        data_start = header_row + 1
        # Buscar columna de inicio del análisis (SCORE TOTAL)
        start_col = None
        for i, h in enumerate(headers):
            if "SCORE" in (h or "").upper() and "TOTAL" in (h or "").upper():
                start_col = i
                break
        if start_col is None:
            _log(f"  ⚠  {sheet_name}: no se encontró columna SCORE TOTAL, se omite")
            continue

        count = 0
        prev_by_sheet[sheet_name] = []
        for row_idx in range(data_start, ws.max_row + 1):
            row = ws[row_idx]
            raw_id = row[id_col].value
            hu_id = str(raw_id or "").strip()
            if hu_id in SKIP_VALUES or not hu_id:
                continue
            norm_id = _normalize_hu_id(hu_id)
            if not norm_id:
                continue
            try:
                title = str(row[title_col].value or "").strip()
            except (IndexError, AttributeError, TypeError):
                title = ""
            try:
                desc = str(row[desc_col].value or "").strip()
            except (IndexError, AttributeError, TypeError):
                desc = ""
            norm_title = _normalize_title(title)
            norm_desc = _normalize_content(desc)

            try:
                score_tot = float(row[start_col].value or 0)
            except (ValueError, TypeError):
                score_tot = 0
            nivel = str(row[start_col + 1].value or "")
            scores = {
                "funcional": _safe_float(row[start_col + 2]),
                "capas_tec": _safe_float(row[start_col + 3]),
                "ux_ui": _safe_float(row[start_col + 4]),
                "integraciones": _safe_float(row[start_col + 5]),
                "regulatorio": _safe_float(row[start_col + 6]),
                "criterios": _safe_float(row[start_col + 7]),
            }
            resumen = str(row[start_col + 10].value or "")
            brechas = {
                "funcional": str(row[start_col + 11].value or ""),
                "capas_tec": str(row[start_col + 12].value or ""),
                "ux_ui": str(row[start_col + 13].value or ""),
                "integraciones": str(row[start_col + 14].value or ""),
                "regulatorio": str(row[start_col + 15].value or ""),
                "criterios": str(row[start_col + 16].value or ""),
            }
            prev_data = {
                "score_total": score_tot,
                "nivel": nivel,
                "scores": scores,
                "resumen": resumen,
                "brechas": brechas,
            }
            prev_by_id[(sheet_name, norm_id)] = prev_data
            prev_by_sheet[sheet_name].append({
                "prev": prev_data,
                "norm_title": norm_title,
                "norm_desc": norm_desc,
            })
            count += 1
        _log(f"  📜  {sheet_name}: {count} HUs de referencia cargadas")
    wb.close()
    return {"by_id": prev_by_id, "by_sheet": prev_by_sheet}


def _find_prev_data(hu: dict, prev_index: dict) -> dict | None:
    """
    Busca el análisis previo para una HU.
    Orden: 1) match por ID, 2) match por título exacto, 3) match por título + contenido similar.
    """
    if not prev_index:
        return None
    by_id = prev_index.get("by_id", {})
    by_sheet = prev_index.get("by_sheet", {})

    sheet = hu["_sheet"]
    norm_id = _normalize_hu_id(hu["_hu_id"])
    title = hu.get("Titulo") or hu.get("Titulo ", "")
    norm_title = _normalize_title(title)
    desc = (hu.get("Descripción") or hu.get("Descripcion") or hu.get("Definición") or hu.get("Definicion") or "")
    norm_desc = _normalize_content(desc)

    # 1. Match por ID
    prev = by_id.get((sheet, norm_id))
    if prev:
        return prev

    # 2. Match por título exacto
    candidates = by_sheet.get(sheet, [])
    if norm_title:
        for c in candidates:
            if c["norm_title"] == norm_title:
                return c["prev"]

    # 3. Match por título + contenido similar (score combinado)
    if not norm_title and not norm_desc:
        return None
    best = None
    best_score = 0.0
    for c in candidates:
        title_sim = _similarity(norm_title, c["norm_title"]) if norm_title and c["norm_title"] else 0
        desc_sim = _similarity(norm_desc, c["norm_desc"]) if norm_desc and c["norm_desc"] else 0
        # Peso: título 60%, contenido 40%
        score = 0.6 * title_sim + 0.4 * desc_sim
        if score > best_score and score >= 0.70:
            best_score = score
            best = c["prev"]
    return best


def _safe_float(v):
    try:
        s = str(v or "").replace("/10", "").strip()
        return float(s) if s else 0
    except (ValueError, TypeError):
        return 0


# ══════════════════════════════════════════════════════════════════════════════
# 5. ANÁLISIS CON CLAUDE
# ══════════════════════════════════════════════════════════════════════════════

def score_to_level(score: float) -> str:
    if score >= 90: return "🟢 Excelente"
    if score >= 75: return "🔵 Completo"
    if score >= 55: return "🟡 Aceptable"
    if score >= 30: return "🟠 En progreso"
    return "🔴 Por definir"


def compute_total_score(scores: dict) -> float:
    """Score total ponderado: convierte scores 0-10 a 0-100."""
    total = sum(
        scores.get(dim, 0) * 10 * weight
        for dim, weight in DIMENSION_WEIGHTS.items()
    )
    return round(total, 1)


def analyze_hu(client: anthropic.Anthropic, hu: dict, prev_data: dict = None, retries: int = 3) -> dict:
    """Envía una HU a Claude y retorna el análisis estructurado."""
    for attempt in range(retries):
        try:
            msg = client.messages.create(
                model="claude-sonnet-4-20250514",
                max_tokens=1800,
                system=SYSTEM_PROMPT,
                messages=[{"role": "user", "content": build_analysis_prompt(hu, prev_data)}]
            )
            raw = msg.content[0].text.strip()

            # Limpiar bloques markdown si los hay
            if raw.startswith("```"):
                lines = raw.split("\n")
                raw = "\n".join(lines[1:-1] if lines[-1].strip() == "```" else lines[1:])

            result = json.loads(raw)
            scores = result.get("scores", {})
            result["score_total"] = compute_total_score(scores)
            result["nivel"] = score_to_level(result["score_total"])
            result.setdefault("mejoras_identificadas", "N/A")
            result.setdefault("comparacion_anterior", "N/A")
            return result

        except json.JSONDecodeError as e:
            if attempt < retries - 1:
                print(f"    ↻  JSON inválido, reintento {attempt + 2}/{retries}...")
                time.sleep(2)
            else:
                return _error_result(f"JSON inválido: {e}")
        except anthropic.RateLimitError:
            wait = 30 * (attempt + 1)
            print(f"    ⏳  Rate limit, esperando {wait}s...")
            time.sleep(wait)
        except Exception as e:
            err_msg = str(e)
            if _is_credits_or_tokens_error(err_msg):
                raise AnthropicGameOverError(err_msg)
            if attempt < retries - 1:
                time.sleep(5)
            else:
                return _error_result(err_msg)

    return _error_result("Máximo de reintentos alcanzado")


def _is_credits_or_tokens_error(msg: str) -> bool:
    """Detecta si el error es por créditos o tokens agotados."""
    m = (msg or "").lower()
    return any(
        x in m for x in [
            "insufficient credits", "credit balance", "too low",
            "credit", "tokens", "blocked", "purchase credits",
            "upgrade", "billing", "spend limit", "usage is blocked"
        ]
    )


def _error_result(msg: str) -> dict:
    return {
        "score_total": 0, "nivel": "⛔ Error",
        "scores": {d: 0 for d in DIMENSIONS},
        "capas_tecnologicas": "No disponible",
        "resumen": f"Error en análisis: {msg}",
        "brechas": {d: "Error en análisis" for d in DIMENSIONS},
        "preguntas_criticas": "No disponible",
        "mejoras_identificadas": "N/A",
        "comparacion_anterior": "N/A",
    }


def _build_initiative_summary_for_executive(sheet_name: str, results: list[dict]) -> str:
    """Construye un resumen compacto de una iniciativa para el prompt de análisis ejecutivo."""
    lines = [f"INICIATIVA: {sheet_name}", f"HUs: {len(results)}, Score promedio: {sum(r['score_total'] for r in results) / len(results):.1f}"]
    for r in results:
        hu_id = r.get("_hu_id", "?")
        score = r.get("score_total", 0)
        nivel = r.get("nivel", "")
        resumen = (r.get("resumen") or "")[:150]
        mejoras = (r.get("mejoras_identificadas") or "N/A")[:120]
        comparacion = (r.get("comparacion_anterior") or "N/A")[:120]
        brechas = r.get("brechas") or {}
        brechas_txt = " | ".join(f"{k}: {(str(v)[:50]+'...' if len(str(v))>50 else v)}" for k, v in brechas.items() if v and "Completo" not in str(v))[:200]
        lines.append(f"  {hu_id} ({score:.0f}, {nivel}): Resumen: {resumen}. Mejoras: {mejoras}. Comparación: {comparacion}. Brechas: {brechas_txt}")
    return "\n".join(lines)


def generate_executive_analysis(client: anthropic.Anthropic, by_sheet: dict[str, list[dict]], silent: bool = False) -> dict[str, str]:
    """
    Genera un párrafo de análisis ejecutivo por iniciativa.
    Retorna dict[nombre_iniciativa] -> párrafo.
    """
    if not by_sheet:
        return {}

    def _log(msg):
        if not silent:
            print(msg)

    summaries = []
    for sheet_name, results in by_sheet.items():
        valid = [r for r in results if r.get("score_total", 0) > 0]
        if not valid:
            continue
        summaries.append(_build_initiative_summary_for_executive(sheet_name, valid))

    if not summaries:
        return {}

    prompt = "Resumen del análisis de HUs por iniciativa:\n\n" + "\n\n---\n\n".join(summaries)

    try:
        msg = client.messages.create(
            model="claude-sonnet-4-20250514",
            max_tokens=4000,
            system=EXECUTIVE_ANALYSIS_PROMPT,
            messages=[{"role": "user", "content": prompt}],
        )
        raw = msg.content[0].text.strip()
        if raw.startswith("```"):
            lines = raw.split("\n")
            raw = "\n".join(lines[1:-1] if lines[-1].strip() == "```" else lines[1:])
        data = json.loads(raw)
        out = {}
        for item in data.get("iniciativas", []):
            nombre = item.get("nombre", "").strip()
            analisis = item.get("analisis_ejecutivo", "").strip()
            if nombre and analisis:
                out[nombre] = analisis
        return out
    except Exception as e:
        _log(f"  ⚠  No se pudo generar análisis ejecutivo: {e}")
        return {}


# ══════════════════════════════════════════════════════════════════════════════
# 6. HELPERS DE ESTILOS EXCEL
# ══════════════════════════════════════════════════════════════════════════════

def _thin_border():
    s = Side(style="thin", color="D0D0D0")
    return Border(left=s, right=s, top=s, bottom=s)

def _thick_border():
    s = Side(style="medium", color="1F3864")
    return Border(left=s, right=s, top=s, bottom=s)

def _h1(cell, text, bg="1F3864"):
    cell.value = text
    cell.font = Font(bold=True, color="FFFFFF", name="Arial", size=12)
    cell.fill = PatternFill("solid", start_color=bg, end_color=bg)
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border = _thick_border()

def _h2(cell, text, bg="2E75B6"):
    cell.value = text
    cell.font = Font(bold=True, color="FFFFFF", name="Arial", size=10)
    cell.fill = PatternFill("solid", start_color=bg, end_color=bg)
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border = _thin_border()

def _data(cell, value, bold=False, bg=None, fc="000000", align="left", size=9):
    cell.value = value
    cell.font = Font(bold=bold, name="Arial", size=size, color=fc)
    cell.alignment = Alignment(horizontal=align, vertical="center", wrap_text=True)
    cell.border = _thin_border()
    if bg:
        cell.fill = PatternFill("solid", start_color=bg, end_color=bg)

def _score_dim_color(score_0_10: float) -> tuple[str, str]:
    """(bg, fc) para scores de dimensión 0-10."""
    s = score_0_10 * 10
    if s >= 90: return "E2EFDA", "375623"
    if s >= 75: return "C6EFCE", "375623"
    if s >= 55: return "FFEB9C", "9C6500"
    if s >= 30: return "FCEBD5", "843C0C"
    return "FFC7CE", "9C0006"

def _score_total_color(score: float) -> tuple[str, str]:
    """(fc, bg) para score total 0-100."""
    if score >= 90: return "375623", "E2EFDA"
    if score >= 75: return "375623", "C6EFCE"
    if score >= 55: return "9C6500", "FFEB9C"
    if score >= 30: return "843C0C", "FCEBD5"
    return "9C0006", "FFC7CE"

def _nivel_color(nivel: str) -> tuple[str, str]:
    """(bg, fc) para el nivel de completitud."""
    m = {
        "🟢 Excelente": ("E2EFDA", "375623"),
        "🔵 Completo":  ("DDEEFF", "1F3864"),
        "🟡 Aceptable": ("FFEB9C", "9C6500"),
        "🟠 En progreso": ("FCEBD5", "843C0C"),
        "🔴 Por definir": ("FFC7CE", "9C0006"),
        "⛔ Error":     ("F2F2F2", "595959"),
    }
    return m.get(nivel, ("F2F2F2", "595959"))

def _fmt_brechas(text: str) -> str:
    if not text or text.strip() in ("Completo", "", "nan"):
        return "✅ Listo para prerefinamiento"
    return "\n".join(f"• {p.strip()}" for p in text.split("|") if p.strip())

def _fmt_capas(text: str) -> str:
    if not text or text.strip() in ("", "nan", "No disponible"):
        return ""
    return "\n".join(f"▸ {p.strip()}" for p in text.split("|") if p.strip())

def _fmt_preguntas(text: str) -> str:
    if not text or text.strip() in ("", "nan", "No disponible"):
        return ""
    return "\n".join(f"❓ {p.strip()}" for p in text.split("|") if p.strip())


def _fmt_mejoras(text: str) -> str:
    if not text or text.strip() in ("", "nan", "N/A"):
        return "N/A"
    return "\n".join(f"✓ {p.strip()}" for p in text.split("|") if p.strip())


def _fmt_comparacion(text: str) -> str:
    if not text or text.strip() in ("", "nan", "N/A"):
        return "N/A"
    return str(text).strip()


# ══════════════════════════════════════════════════════════════════════════════
# 7. ESCRITURA DE ANÁLISIS EN HOJAS DE HUs
# ══════════════════════════════════════════════════════════════════════════════

def write_analysis_to_sheet(ws, results_by_row: dict):
    """Agrega columnas de análisis a la derecha de los datos existentes."""
    header_row = _detect_header_row(ws)
    last_col = ws.max_column
    sep_col  = last_col + 1
    start_col = sep_col + 1

    # Columna separadora visual azul
    ws.column_dimensions[get_column_letter(sep_col)].width = 2
    for row in range(1, ws.max_row + 1):
        ws.cell(row=row, column=sep_col).fill = PatternFill(
            "solid", start_color="1F3864", end_color="1F3864"
        )

    # Headers de análisis (usa la fila detectada para esta hoja)
    ws.row_dimensions[header_row].height = 50
    for i, (hdr_text, color, width) in enumerate(ANALYSIS_HEADERS):
        col = start_col + i
        _h2(ws.cell(row=header_row, column=col), hdr_text, bg=color)
        ws.column_dimensions[get_column_letter(col)].width = width

    # Datos por fila
    row_fills = ["FFFFFF", "F5F8FF"]
    for row_idx, result in results_by_row.items():
        ws.row_dimensions[row_idx].height = 100
        bg_row = row_fills[row_idx % 2]

        scores    = result.get("scores", {})
        score_tot = result.get("score_total", 0)
        nivel     = result.get("nivel", "⛔ Error")
        brechas   = result.get("brechas", {})

        col_values = [
            score_tot,
            nivel,
            scores.get("funcional", 0),
            scores.get("capas_tec", 0),
            scores.get("ux_ui", 0),
            scores.get("integraciones", 0),
            scores.get("regulatorio", 0),
            scores.get("criterios", 0),
            _fmt_capas(result.get("capas_tecnologicas", "")),
            result.get("resumen", ""),
            _fmt_brechas(brechas.get("funcional", "")),
            _fmt_brechas(brechas.get("capas_tec", "")),
            _fmt_brechas(brechas.get("ux_ui", "")),
            _fmt_brechas(brechas.get("integraciones", "")),
            _fmt_brechas(brechas.get("regulatorio", "")),
            _fmt_brechas(brechas.get("criterios", "")),
            _fmt_preguntas(result.get("preguntas_criticas", "")),
            _fmt_mejoras(result.get("mejoras_identificadas", "N/A")),
            _fmt_comparacion(result.get("comparacion_anterior", "N/A")),
        ]

        for i, value in enumerate(col_values):
            col = start_col + i
            cell = ws.cell(row=row_idx, column=col, value=value)
            cell.font = Font(name="Arial", size=9)
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            cell.border = _thin_border()
            cell.fill = PatternFill("solid", start_color=bg_row, end_color=bg_row)

            if i == 0:  # Score total
                fc, bg = _score_total_color(score_tot)
                cell.font = Font(bold=True, name="Arial", size=14, color=fc)
                cell.fill = PatternFill("solid", start_color=bg, end_color=bg)
                cell.alignment = Alignment(horizontal="center", vertical="center")

            elif i == 1:  # Nivel
                bg, fc = _nivel_color(nivel)
                cell.font = Font(bold=True, name="Arial", size=9, color=fc)
                cell.fill = PatternFill("solid", start_color=bg, end_color=bg)
                cell.alignment = Alignment(horizontal="center", vertical="center")

            elif 2 <= i <= 7:  # Scores por dimensión
                bg, fc = _score_dim_color(float(value))
                cell.font = Font(bold=True, name="Arial", size=10, color=fc)
                cell.fill = PatternFill("solid", start_color=bg, end_color=bg)
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.value = f"{value}/10"

            elif i == 8:  # Capas tecnológicas involucradas
                cell.font = Font(name="Arial", size=9, color="1F3864")
                cell.fill = PatternFill("solid", start_color="E8F4FD", end_color="E8F4FD")

            elif 10 <= i <= 15:  # Brechas
                if str(value).startswith("✅"):
                    cell.font = Font(name="Arial", size=9, color="375623")
                    cell.fill = PatternFill("solid", start_color="E2EFDA", end_color="E2EFDA")
                else:
                    cell.fill = PatternFill("solid", start_color="FFF5F5", end_color="FFF5F5")

            elif i == 16:  # Preguntas para prerefinamiento
                cell.font = Font(bold=True, name="Arial", size=9, color="7B2C2C")
                cell.fill = PatternFill("solid", start_color="FFF0F0", end_color="FFF0F0")
            elif i == 17:  # Mejoras identificadas
                cell.font = Font(name="Arial", size=9, color="375623")
                cell.fill = PatternFill("solid", start_color="E2EFDA", end_color="E2EFDA")
            elif i == 18:  # Comparación vs anterior
                cell.font = Font(name="Arial", size=9, color="7B2C2C")
                cell.fill = PatternFill("solid", start_color="FFF5F5", end_color="FFF5F5")


# ══════════════════════════════════════════════════════════════════════════════
# 8. HOJA DE SÍNTESIS EJECUTIVA
# ══════════════════════════════════════════════════════════════════════════════

def create_synthesis_sheet(wb, all_results: list[dict]):
    """Crea la hoja '📊 Síntesis Ejecutiva' con KPIs, tabla de scores y brechas."""

    if "📊 Síntesis Ejecutiva" in wb.sheetnames:
        del wb["📊 Síntesis Ejecutiva"]
    ws = wb.create_sheet("📊 Síntesis Ejecutiva", 0)

    # Anchos de columna fijos
    col_widths = {
        "A": 5, "B": 24, "C": 14,
        "D": 13, "E": 13, "F": 13,
        "G": 13, "H": 13, "I": 13,
        "J": 2,  "K": 14, "L": 3,
        "M": 55,
    }
    for col, w in col_widths.items():
        ws.column_dimensions[col].width = w

    valid = [r for r in all_results if r.get("score_total", 0) > 0]
    if not valid:
        ws["A1"].value = "Sin datos válidos para mostrar."
        return

    # Agrupar por iniciativa (hoja)
    by_sheet: dict[str, list] = {}
    for r in valid:
        by_sheet.setdefault(r.get("_sheet", "Sin hoja"), []).append(r)

    cur = 1  # cursor de fila actual

    # ── TÍTULO ────────────────────────────────────────────────────────────
    ws.row_dimensions[cur].height = 40
    ws.merge_cells(f"A{cur}:M{cur}")
    _h1(ws[f"A{cur}"],
        "📊  SÍNTESIS EJECUTIVA — DEFINICIÓN FUNCIONAL + CAPAS TECNOLÓGICAS  ·  ACTINVER")
    cur += 1

    ws.row_dimensions[cur].height = 16
    ws.merge_cells(f"A{cur}:M{cur}")
    ws[f"A{cur}"].value = (
        f"Generado: {datetime.now().strftime('%d/%m/%Y %H:%M')}  |  "
        f"HUs analizadas: {len(valid)}  |  "
        f"Iniciativas: {', '.join(by_sheet.keys())}"
    )
    ws[f"A{cur}"].font = Font(italic=True, name="Arial", size=9, color="595959")
    ws[f"A{cur}"].alignment = Alignment(horizontal="center", vertical="center")
    cur += 2

    # ══════════════════════════════════════════════════════════════════════
    # SECCIÓN A: KPIs GLOBALES
    # ══════════════════════════════════════════════════════════════════════
    ws.row_dimensions[cur].height = 25
    ws.merge_cells(f"A{cur}:M{cur}")
    _h1(ws[f"A{cur}"], "▌ A.  MÉTRICAS GLOBALES")
    cur += 1

    all_scores = [r["score_total"] for r in valid]
    nivel_counts: dict[str, int] = {}
    for r in valid:
        n = r.get("nivel", "")
        nivel_counts[n] = nivel_counts.get(n, 0) + 1

    avg_global = sum(all_scores) / len(all_scores)

    kpi_data = [
        ("Score Promedio Global",     f"{avg_global:.1f} / 100",                         "2E75B6"),
        ("Score Máximo",              f"{max(all_scores):.0f} / 100",                    "375623"),
        ("Score Mínimo",              f"{min(all_scores):.0f} / 100",                    "843C0C"),
        ("🟢  Excelentes  (90-100)", f"{nivel_counts.get('🟢 Excelente', 0)} HUs",      "375623"),
        ("🔵  Completas   (75-89)",  f"{nivel_counts.get('🔵 Completo', 0)} HUs",       "1F3864"),
        ("🟡  Aceptables  (55-74)",  f"{nivel_counts.get('🟡 Aceptable', 0)} HUs",      "9C6500"),
        ("🟠  En progreso (30-54)",  f"{nivel_counts.get('🟠 En progreso', 0)} HUs",   "843C0C"),
        ("🔴  Por definir (0-29)",   f"{nivel_counts.get('🔴 Por definir', 0)} HUs",    "9C0006"),
    ]

    for label, value, fc in kpi_data:
        ws.row_dimensions[cur].height = 18
        ws.merge_cells(f"A{cur}:D{cur}")
        _data(ws[f"A{cur}"], label, bold=True, fc=fc, bg="F7F9FF")
        ws.merge_cells(f"E{cur}:G{cur}")
        _data(ws[f"E{cur}"], value, bold=True, align="center", fc=fc, bg="F7F9FF")
        ws.merge_cells(f"H{cur}:M{cur}")
        cur += 1

    cur += 1

    # ══════════════════════════════════════════════════════════════════════
    # SECCIÓN B: TABLA DE SCORES POR INICIATIVA Y DIMENSIÓN
    # ══════════════════════════════════════════════════════════════════════
    ws.row_dimensions[cur].height = 25
    ws.merge_cells(f"A{cur}:M{cur}")
    _h1(ws[f"A{cur}"], "▌ B.  SCORES POR INICIATIVA Y DIMENSIÓN")
    cur += 1

    # Sub-encabezados de tabla (Definición funcional + capas tecnológicas)
    ws.row_dimensions[cur].height = 50
    sub_hdrs = [
        ("A", "#",                          "44546A"),
        ("B", "INICIATIVA\n(Producto)",     "44546A"),
        ("C", "SCORE TOTAL\n(0-100)",       "1F3864"),
        ("D", "Funcional\n(0-10)\n35%",    "2E75B6"),
        ("E", "Capas Tec.\n(0-10)\n25%",   "2E75B6"),
        ("F", "UX/UI\n(0-10)\n15%",       "2E75B6"),
        ("G", "Integr.\n(0-10)\n10%",     "2E75B6"),
        ("H", "Regulat.\n(0-10)\n8%",     "2E75B6"),
        ("I", "Criterios\n(0-10)\n7%",    "2E75B6"),
        ("J", "",                           "FFFFFF"),
        ("K", "NIVEL GENERAL",             "44546A"),
        ("L", "",                           "FFFFFF"),
        ("M", "HUs CON MAYOR OPORTUNIDAD DE MEJORA",  "7B2C2C"),
    ]
    for col_l, txt, bg in sub_hdrs:
        _h2(ws[f"{col_l}{cur}"], txt, bg=bg)
    cur += 1

    row_fills = ["FFFFFF", "F5F8FF"]
    table_data_start = cur

    dim_cols = [
        ("funcional", "D"), ("capas_tec", "E"), ("ux_ui", "F"),
        ("integraciones", "G"), ("regulatorio", "H"), ("criterios", "I"),
    ]

    for sheet_idx, (sheet_name, sheet_results) in enumerate(by_sheet.items(), 1):
        ws.row_dimensions[cur].height = 22
        fill = row_fills[cur % 2]

        avg_total = sum(r["score_total"] for r in sheet_results) / len(sheet_results)
        avg_dims  = {
            d: sum(r.get("scores", {}).get(d, 0) for r in sheet_results) / len(sheet_results)
            for d in DIMENSIONS
        }
        nivel = score_to_level(avg_total)

        # Contar HUs por definir y en progreso
        por_definir = sum(1 for r in sheet_results if r["score_total"] < 30)
        en_progreso = sum(1 for r in sheet_results if 30 <= r["score_total"] < 55)
        criticas_txt = ""
        if por_definir:
            ids_pd = [r["_hu_id"] for r in sheet_results if r["score_total"] < 30]
            criticas_txt += f"🔴 Por definir: {', '.join(ids_pd)}\n"
        if en_progreso:
            ids_ep = [r["_hu_id"] for r in sheet_results if 30 <= r["score_total"] < 55]
            criticas_txt += f"🟠 En progreso: {', '.join(ids_ep)}"
        if not criticas_txt:
            criticas_txt = "✅ Ninguna"

        _data(ws[f"A{cur}"], sheet_idx, align="center", bg=fill)
        _data(ws[f"B{cur}"], f"{sheet_name}  ({len(sheet_results)} HUs)",
              bold=True, bg=fill)

        fc_t, bg_t = _score_total_color(avg_total)
        _data(ws[f"C{cur}"], round(avg_total, 1), bold=True, align="center",
              fc=fc_t, bg=bg_t, size=12)

        for dim, col in dim_cols:
            val = round(avg_dims[dim], 1)
            bg_d, fc_d = _score_dim_color(val)
            _data(ws[f"{col}{cur}"], f"{val}/10", bold=True, align="center",
                  fc=fc_d, bg=bg_d)

        ws[f"J{cur}"].fill = PatternFill("solid", start_color="1F3864", end_color="1F3864")

        bg_n, fc_n = _nivel_color(nivel)
        _data(ws[f"K{cur}"], nivel, bold=True, align="center", fc=fc_n, bg=bg_n)

        ws[f"L{cur}"].fill = PatternFill("solid", start_color="1F3864", end_color="1F3864")

        cell_crit = ws[f"M{cur}"]
        cell_crit.value = criticas_txt
        cell_crit.font = Font(name="Arial", size=9,
                              color="375623" if criticas_txt == "✅ Ninguna" else "7B2C2C")
        cell_crit.fill = PatternFill("solid",
                                     start_color="E2EFDA" if criticas_txt == "✅ Ninguna" else "FFF5F5",
                                     end_color="E2EFDA" if criticas_txt == "✅ Ninguna" else "FFF5F5")
        cell_crit.alignment = Alignment(vertical="top", wrap_text=True)
        cell_crit.border = _thin_border()
        cur += 1

    # Fila de PROMEDIO GENERAL
    ws.row_dimensions[cur].height = 28
    avg_grand = sum(all_scores) / len(all_scores)
    avg_grand_dims = {
        d: sum(r.get("scores", {}).get(d, 0) for r in valid) / len(valid)
        for d in DIMENSIONS
    }
    ws.merge_cells(f"A{cur}:B{cur}")
    _data(ws[f"A{cur}"], "▶  PROMEDIO GENERAL", bold=True,
          bg="1F3864", fc="FFFFFF", align="center", size=10)

    fc_g, bg_g = _score_total_color(avg_grand)
    _data(ws[f"C{cur}"], round(avg_grand, 1), bold=True, align="center",
          fc=fc_g, bg=bg_g, size=13)

    for dim, col in dim_cols:
        val = round(avg_grand_dims[dim], 1)
        bg_d, fc_d = _score_dim_color(val)
        _data(ws[f"{col}{cur}"], f"{val}/10", bold=True, align="center",
              fc=fc_d, bg=bg_d)

    ws[f"J{cur}"].fill = PatternFill("solid", start_color="1F3864", end_color="1F3864")
    bg_n, fc_n = _nivel_color(score_to_level(avg_grand))
    _data(ws[f"K{cur}"], score_to_level(avg_grand), bold=True, align="center",
          fc=fc_n, bg=bg_n)
    ws[f"L{cur}"].fill = PatternFill("solid", start_color="1F3864", end_color="1F3864")
    cur += 2

    # ══════════════════════════════════════════════════════════════════════
    # SECCIÓN C: BRECHAS CONSOLIDADAS POR INICIATIVA
    # ══════════════════════════════════════════════════════════════════════
    ws.row_dimensions[cur].height = 25
    ws.merge_cells(f"A{cur}:M{cur}")
    _h1(ws[f"A{cur}"], "▌ C.  ELEMENTOS POR DEFINIR POR INICIATIVA", bg="843C0C")
    cur += 1

    # Sub-encabezados
    ws.row_dimensions[cur].height = 40
    br_hdrs = [
        ("A", "#",                       "44546A"),
        ("B", "INICIATIVA",              "44546A"),
        ("C", "DIM. A\nFORTALECER",    "843C0C"),
        ("D", "SCORE",                  "843C0C"),
        ("E", "2ª DIM.\nA FORTALECER",  "C0392B"),
        ("F", "SCORE",                  "C0392B"),
    ]
    for col_l, txt, bg in br_hdrs:
        _h2(ws[f"{col_l}{cur}"], txt, bg=bg)

    ws.merge_cells(f"G{cur}:M{cur}")
    _h2(ws[f"G{cur}"],
        "ELEMENTOS POR DEFINIR MÁS FRECUENTES  (número = HUs que lo requieren)",
        bg="843C0C")
    cur += 1

    for sheet_idx, (sheet_name, sheet_results) in enumerate(by_sheet.items(), 1):
        ws.row_dimensions[cur].height = 110
        fill = row_fills[cur % 2]

        avg_dims = {
            d: sum(r.get("scores", {}).get(d, 0) for r in sheet_results) / len(sheet_results)
            for d in DIMENSIONS
        }

        # Las dos dimensiones más débiles
        sorted_dims = sorted(avg_dims.items(), key=lambda x: x[1])
        weakest1_dim, weakest1_score = sorted_dims[0]
        weakest2_dim, weakest2_score = sorted_dims[1] if len(sorted_dims) > 1 else (sorted_dims[0])

        # Contar frecuencia de brechas en TODAS las dimensiones
        brecha_freq: dict[str, int] = {}
        for r in sheet_results:
            for d in DIMENSIONS:
                raw = r.get("brechas", {}).get(d, "")
                if raw and raw not in ("Completo", "", "nan"):
                    for item in raw.split("|"):
                        item = item.strip()
                        if len(item) > 12:
                            # Clave de agrupación: primeras 70 chars
                            key = item[:70]
                            brecha_freq[key] = brecha_freq.get(key, 0) + 1

        top_brechas = sorted(brecha_freq.items(), key=lambda x: x[1], reverse=True)[:6]
        brechas_txt = "\n".join(
            f"[{cnt}]  {brecha}" for brecha, cnt in top_brechas
        ) if top_brechas else "✅ Todas las HUs tienen definición suficiente en estas dimensiones"

        _data(ws[f"A{cur}"], sheet_idx, align="center", bg=fill)
        _data(ws[f"B{cur}"], f"{sheet_name}\n({len(sheet_results)} HUs)",
              bold=True, bg=fill)

        bg_w1, fc_w1 = _score_dim_color(weakest1_score)
        _data(ws[f"C{cur}"], DIMENSIONS[weakest1_dim], bold=True, fc=fc_w1, bg=bg_w1)
        _data(ws[f"D{cur}"], f"{weakest1_score:.1f}/10", bold=True,
              align="center", fc=fc_w1, bg=bg_w1)

        bg_w2, fc_w2 = _score_dim_color(weakest2_score)
        _data(ws[f"E{cur}"], DIMENSIONS[weakest2_dim], bold=True, fc=fc_w2, bg=bg_w2)
        _data(ws[f"F{cur}"], f"{weakest2_score:.1f}/10", bold=True,
              align="center", fc=fc_w2, bg=bg_w2)

        ws.merge_cells(f"G{cur}:M{cur}")
        cell = ws[f"G{cur}"]
        cell.value = brechas_txt
        is_clean = brechas_txt.startswith("✅")
        cell.font = Font(name="Arial", size=9,
                         color="375623" if is_clean else "7B2C2C")
        cell.fill = PatternFill("solid",
                                start_color="E2EFDA" if is_clean else "FFF5F5",
                                end_color="E2EFDA" if is_clean else "FFF5F5")
        cell.alignment = Alignment(vertical="top", wrap_text=True)
        cell.border = _thin_border()
        cur += 1

    cur += 2

    # ══════════════════════════════════════════════════════════════════════
    # SECCIÓN D: LEYENDA
    # ══════════════════════════════════════════════════════════════════════
    ws.row_dimensions[cur].height = 25
    ws.merge_cells(f"A{cur}:M{cur}")
    _h1(ws[f"A{cur}"], "▌ D.  LEYENDA DE SCORING", bg="44546A")
    cur += 1

    leyenda = [
        ("🟢 Excelente  90–100", "Lista para prerefinamiento.",                          "E2EFDA", "375623"),
        ("🔵 Completa   75–89",  "Lista con pequeñas clarificaciones opcionales.",        "DDEEFF", "1F3864"),
        ("🟡 Aceptable  55–74",  "Conviene definir algunos elementos antes del prerefinamiento.", "FFEB9C", "9C6500"),
        ("🟠 En progreso 30–54", "El PO puede fortalecer la definición en varias dimensiones.", "FCEBD5", "843C0C"),
        ("🔴 Por definir 0–29",  "Oportunidad de definir más antes de involucrar a técnicos.",  "FFC7CE", "9C0006"),
    ]
    for nivel_txt, desc, bg, fc in leyenda:
        ws.row_dimensions[cur].height = 18
        ws.merge_cells(f"A{cur}:D{cur}")
        _data(ws[f"A{cur}"], nivel_txt, bold=True, fc=fc, bg=bg)
        ws.merge_cells(f"E{cur}:M{cur}")
        _data(ws[f"E{cur}"], desc, fc=fc, bg=bg)
        cur += 1

    cur += 1
    ws.row_dimensions[cur].height = 22
    ws.merge_cells(f"A{cur}:M{cur}")
    weights_txt = "  |  ".join(
        f"{DIMENSIONS[d]}: {int(w*100)}%"
        for d, w in DIMENSION_WEIGHTS.items()
    )
    _data(ws[f"A{cur}"],
          f"Pesos del score total ponderado:  {weights_txt}",
          bold=True, fc="1F3864", bg="EBF5FF")

    ws.freeze_panes = f"C5"


# ══════════════════════════════════════════════════════════════════════════════
# 9. ORQUESTADOR PRINCIPAL
# ══════════════════════════════════════════════════════════════════════════════

def run(input_path: str, output_path: str,
        target_sheet: str = None, limit: int = None, silent: bool = False,
        previous_analysis_path: str = None,
        progress_callback=None) -> dict:
    """
    Ejecuta el análisis de HUs. Retorna un dict con el resumen para uso programático.
    """
    def log(msg=""):
        if not silent:
            print(msg)

    log("\n" + "═" * 60)
    log("  HU ANALYZER — Actinver Digital Products")
    log("═" * 60)

    try:
        from config import get_api_key
        api_key = get_api_key()
    except ImportError:
        api_key = os.environ.get("ANTHROPIC_API_KEY", "")
    if not api_key:
        err = "ANTHROPIC_API_KEY no configurada."
        if not silent:
            print(f"\n❌  {err}")
            print("    Windows:   set ANTHROPIC_API_KEY=sk-ant-...")
            print("    Mac/Linux: export ANTHROPIC_API_KEY=sk-ant-...")
            sys.exit(1)
        raise ValueError(err)

    client = anthropic.Anthropic(api_key=api_key)

    log(f"\n📂  Cargando: {input_path}")
    _, all_hus = load_all_hus(input_path, target_sheet, quiet=silent)
    total = len(all_hus)

    prev_index = {}
    if previous_analysis_path and os.path.exists(previous_analysis_path):
        try:
            prev_index = load_previous_analysis(previous_analysis_path, log_fn=log)
            total_prev = len(prev_index.get("by_id", {}))
            log(f"  📜  Total análisis anterior: {total_prev} HUs de referencia (match por ID, título y contenido)")
        except Exception as e:
            if not silent:
                print(f"  ⚠  No se pudo cargar análisis anterior: {e}")

    if limit:
        all_hus = all_hus[:limit]
        log(f"\n  ⚙  Modo prueba: {len(all_hus)} de {total} HUs")

    workers = min(MAX_CONCURRENT_ANALYSIS, len(all_hus))
    total_hus = len(all_hus)
    log(f"\n🔍  Analizando {total_hus} HUs con Claude AI ({workers} en paralelo)...\n")

    hu_speed = get_hu_speed()
    if progress_callback:
        eta = (total_hus * (hu_speed or 15)) / workers if hu_speed else None
        progress_callback(0, total_hus, hu_speed or 0, 0, eta or 0)

    def _analyze_one(args):
        client_ref, idx, hu, prev = args
        result = analyze_hu(client_ref, hu, prev_data=prev)
        result["_sheet"] = hu["_sheet"]
        result["_row"] = hu["_row"]
        result["_hu_id"] = hu["_hu_id"]
        return (idx, result)

    results_by_idx: dict[int, dict] = {}
    completed = 0
    future_to_start: dict = {}

    with ThreadPoolExecutor(max_workers=workers) as executor:
        for idx, hu in enumerate(all_hus, 1):
            prev = _find_prev_data(hu, prev_index)
            fut = executor.submit(_analyze_one, (client, idx, hu, prev))
            future_to_start[fut] = (idx, hu, time.time())

        for future in as_completed(future_to_start):
            idx, hu, start_time = future_to_start[future]
            elapsed = time.time() - start_time
            completed += 1

            try:
                _, result = future.result()
                results_by_idx[idx] = result
                score = result.get("score_total", 0)
                nivel = result.get("nivel", "?")
                title = hu.get("Titulo", hu.get("Titulo ", "Sin título"))[:50]
                log(f"  [{completed:3}/{total_hus}]  {hu['_sheet']:20} | {hu['_hu_id']:10} | {title}")
                log(f"             → {nivel}  ({score:.0f}/100)")
            except AnthropicGameOverError:
                raise
            except Exception as e:
                results_by_idx[idx] = _error_result(str(e))
                results_by_idx[idx]["_sheet"] = hu["_sheet"]
                results_by_idx[idx]["_row"] = hu["_row"]
                results_by_idx[idx]["_hu_id"] = hu["_hu_id"]
                log(f"  [{completed:3}/{total_hus}]  {hu['_sheet']:20} | {hu['_hu_id']:10} | ⛔ Error: {e}")

            hu_speed = update_hu_speed(elapsed)
            if progress_callback:
                remaining = total_hus - completed
                eta = (remaining * hu_speed / workers) if hu_speed and remaining > 0 else 0
                progress_callback(completed, total_hus, hu_speed or 0, elapsed, eta)

    results_by_sheet_row: dict[str, dict] = {}
    all_results_flat = [results_by_idx[i] for i in range(1, len(all_hus) + 1)]
    for r in all_results_flat:
        results_by_sheet_row.setdefault(r["_sheet"], {})[r["_row"]] = r

    log(f"\n💾  Escribiendo: {output_path}")
    wb = openpyxl.load_workbook(input_path)

    for sheet_name, row_results in results_by_sheet_row.items():
        if sheet_name not in wb.sheetnames:
            continue
        write_analysis_to_sheet(wb[sheet_name], row_results)
        log(f"  ✓  {sheet_name}: {len(row_results)} HUs")

    log("\n📊  Generando Síntesis Ejecutiva...")
    create_synthesis_sheet(wb, all_results_flat)
    wb.move_sheet("📊 Síntesis Ejecutiva", offset=-(len(wb.sheetnames) - 1))
    wb.save(output_path)

    valid = [r for r in all_results_flat if r.get("score_total", 0) > 0]
    avg = sum(r["score_total"] for r in valid) / len(valid) if valid else 0
    criticas    = sum(1 for r in valid if r["score_total"] < 30)
    incompletas = sum(1 for r in valid if 30 <= r["score_total"] < 55)
    excelentes  = sum(1 for r in valid if r["score_total"] >= 90)
    completas   = sum(1 for r in valid if 75 <= r["score_total"] < 90)
    aceptables  = sum(1 for r in valid if 55 <= r["score_total"] < 75)

    by_sheet = {}
    by_sheet_full: dict[str, list[dict]] = {}
    for r in valid:
        s = r.get("_sheet", "Sin hoja")
        if s not in by_sheet:
            by_sheet[s] = {"count": 0, "avg": 0, "scores": []}
            by_sheet_full[s] = []
        by_sheet[s]["count"] += 1
        by_sheet[s]["scores"].append(r["score_total"])
        by_sheet_full[s].append(r)
    for s, d in by_sheet.items():
        d["avg"] = round(sum(d["scores"]) / len(d["scores"]), 1) if d["scores"] else 0
        del d["scores"]

    log("\n📋  Generando análisis ejecutivo por iniciativa...")
    executive_by_initiative = generate_executive_analysis(client, by_sheet_full, silent=silent)

    log(f"""
{"═"*60}
  ✅  ANÁLISIS COMPLETADO
{"═"*60}
  HUs analizadas : {len(valid)}
  Score promedio : {avg:.1f} / 100
  HUs por definir: {criticas}  (score < 30)
  HUs en progreso : {incompletas}  (score 30–54)
  Archivo output : {output_path}
{"═"*60}
""")

    return {
        "total_hus": len(valid),
        "avg_score": round(avg, 1),
        "criticas": criticas,
        "incompletas": incompletas,
        "aceptables": aceptables,
        "completas": completas,
        "excelentes": excelentes,
        "by_sheet": by_sheet,
        "executive_by_initiative": executive_by_initiative,
        "output_path": output_path,
    }


# ══════════════════════════════════════════════════════════════════════════════
# 10. ENTRY POINT
# ══════════════════════════════════════════════════════════════════════════════

if __name__ == "__main__":
    parser = argparse.ArgumentParser(
        description="Analiza completitud de HUs con Claude AI — Actinver"
    )
    parser.add_argument("--input",  "-i", required=True,
                        help="Excel de entrada (ej: HUs_Compilado.xlsx)")
    parser.add_argument("--output", "-o", default=None,
                        help="Excel de salida (default: Output/input_analizado_vN.0.xlsx)")
    parser.add_argument("--sheet",  "-s", default=None,
                        help="Analizar solo una hoja (ej: 'Onboarding')")
    parser.add_argument("--limit",  "-l", type=int, default=None,
                        help="Limitar a N HUs para pruebas")
    parser.add_argument("--previous", "-p", default=None,
                        help="Excel del análisis anterior para comparar mejoras")

    args = parser.parse_args()

    if not args.output:
        args.output = get_next_output_path(args.input)

    run(args.input, args.output, args.sheet, args.limit, previous_analysis_path=args.previous)
